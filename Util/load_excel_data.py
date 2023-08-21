from openpyxl import load_workbook
import os

location_to_filename = {
    "뚝섬 한강공원": "뚝섬",
    "여의도 한강공원": "여의도",
    "망원 한강공원": "망원",
    "강서 한강공원": "강서",
    "양화 한강공원": "양화",
    "이촌 한강공원": "이촌",
    "반포 한강공원": "반포",
    "잠원 한강공원": "잠원",
    "잠실 한강공원": "잠실",
    "광나루 한강공원": "광나루",
    "난지 한강공원": "난지"
}

def get_cell_name(day):
    # E = 1일 AA = 23
    if day >= 23:
        return 'A'+chr(day - 23 + ord('A'))
    return chr(day - 1 + ord('E'))

def get_data(year, month, day, place):
    file_name = f'C:/_temporary/Hackuthon_back/Util/han_lake/{year}년/{location_to_filename[place]}.xlsx'

    wb = load_workbook(file_name, 
                    read_only=True, ## 읽기 전용(읽기 전용에 최적화되어 파일을 불러온다)
                    data_only=True, ## False면 셀안 공식을 가져오고 True면 공식 적용된 값만을 불러온다.
                    )
    ws = wb[wb.sheetnames[month-1]] # 첫 번째 시트
    day = get_cell_name(day)
    #print(day)
    res = []
    for i in range(5, 9):
        res.append(ws[f'{day}{i}'].value)
    # 5 = 날씨, 6은 아침, 7 점심 8 저녁
    wb.close() ## Workbook 종료
    return res

#get_data(2021, 8, 25, "강서")